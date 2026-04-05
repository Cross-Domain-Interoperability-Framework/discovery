"""Merge all CDIF metadata crosswalk workbooks into one."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

directory = r'C:\Users\smrTu\OneDrive\Documents\GithubC\CDIF\Discovery'

###############################################################################
# 1. Load base: crosswalks draft
###############################################################################
wb_base = openpyxl.load_workbook(
    os.path.join(directory, 'CDIF metadata cross walks.xlsx'), data_only=True)
ws_base = wb_base['crosswalks draft']

base_headers_raw = []
for row in ws_base.iter_rows(min_row=1, max_row=1, values_only=True):
    base_headers_raw = list(row)

base_headers = []
for h in base_headers_raw:
    if h is not None:
        base_headers.append(str(h).strip())
    else:
        break

print(f'Base headers ({len(base_headers)})')

base_rows = []
for row in ws_base.iter_rows(min_row=2, values_only=True):
    vals = list(row[:len(base_headers)])
    while len(vals) < len(base_headers):
        vals.append(None)
    if any(v is not None for v in vals):
        base_rows.append(dict(zip(base_headers, vals)))

print(f'Base rows: {len(base_rows)}')
wb_base.close()

###############################################################################
# 2. Load DDE-software-etc - add missing rows and fill gaps
###############################################################################
wb_dde = openpyxl.load_workbook(
    os.path.join(directory, 'metadataCrosswalkCDIF-DDE-software-etc.xlsx'), data_only=True)
ws_dde = wb_dde['crosswalk']

dde_headers_raw = []
for row in ws_dde.iter_rows(min_row=1, max_row=1, values_only=True):
    dde_headers_raw = [str(h).strip() if h else None for h in row]
dde_headers = [h for h in dde_headers_raw if h]

dde_col_map = {
    'generic metadata item': 'Generic element',
    'CDIF content item': 'CDIF content model',
    'Property': 'schema:Property',
    'Type': 'Domain Type',
    'Dublin Core': 'DC and DC Term',
}

base_by_ge = {}
for r in base_rows:
    ge = str(r.get('Generic element', '')).strip()
    if ge:
        base_by_ge[ge] = r

dde_added = 0
for row in ws_dde.iter_rows(min_row=2, values_only=True):
    vals = list(row[:len(dde_headers)])
    while len(vals) < len(dde_headers):
        vals.append(None)
    dde_row = dict(zip(dde_headers, vals))

    ge = str(dde_row.get('generic metadata item', '')).strip()
    if not ge:
        continue

    if ge in base_by_ge:
        for dde_col, val in dde_row.items():
            if val is None:
                continue
            base_col = dde_col_map.get(dde_col, dde_col)
            if base_col in base_headers:
                if base_by_ge[ge].get(base_col) is None:
                    base_by_ge[ge][base_col] = val
    else:
        new_row = {h: None for h in base_headers}
        for dde_col, val in dde_row.items():
            base_col = dde_col_map.get(dde_col, dde_col)
            if base_col in base_headers:
                new_row[base_col] = val
        base_rows.append(new_row)
        base_by_ge[ge] = new_row
        dde_added += 1

print(f'Added {dde_added} rows from DDE-software-etc')
wb_dde.close()

###############################################################################
# 3. Load DCAT-ISOetc - fill gaps by CDIF content model
###############################################################################
wb_dcat = openpyxl.load_workbook(
    os.path.join(directory, 'metadataCrosswalksCDIF-DCAT-ISOetc.xlsx'), data_only=True)
ws_dcat = wb_dcat['crosswalks']

dcat_headers_raw = []
for row in ws_dcat.iter_rows(min_row=1, max_row=1, values_only=True):
    dcat_headers_raw = [str(h).strip() if h else None for h in row]
dcat_headers = [h for h in dcat_headers_raw if h]

# Normalize EOSC header
for i, h in enumerate(dcat_headers):
    if h and 'EOSC' in h:
        dcat_headers[i] = 'EOSC/EDMI'

base_by_cm = {}
for r in base_rows:
    cm = str(r.get('CDIF content model', '')).strip()
    if cm:
        if cm not in base_by_cm:
            base_by_cm[cm] = []
        base_by_cm[cm].append(r)

dcat_filled = 0
dcat_new = 0
for row in ws_dcat.iter_rows(min_row=2, values_only=True):
    vals = list(row[:len(dcat_headers)])
    while len(vals) < len(dcat_headers):
        vals.append(None)
    dcat_row = dict(zip(dcat_headers, vals))

    cm = str(dcat_row.get('CDIF content model', '')).strip()
    ge = str(dcat_row.get('Generic element', '')).strip()

    targets = base_by_cm.get(cm, [])
    if targets:
        for t in targets:
            for col, val in dcat_row.items():
                if val is None or col == 'Count':
                    continue
                if col in base_headers and t.get(col) is None:
                    t[col] = val
                    dcat_filled += 1
    elif ge:
        if ge not in base_by_ge:
            new_row = {h: None for h in base_headers}
            for col, val in dcat_row.items():
                if col in base_headers and col != 'Count':
                    new_row[col] = val
            base_rows.append(new_row)
            base_by_ge[ge] = new_row
            dcat_new += 1

print(f'Filled {dcat_filled} cells, added {dcat_new} rows from DCAT-ISOetc')
wb_dcat.close()

###############################################################################
# 4. Add implementation columns from 3 smaller sheets
###############################################################################
new_cols = [
    'Obl.', 'Schema.org implementation', 'DCAT v3 implementation',
    'SOSO comparison', 'Scope note (SDO)', 'Scope note (DCAT)', 'Scope note (SOSO)'
]
all_headers = base_headers + new_cols

for r in base_rows:
    for nc in new_cols:
        r[nc] = None

# Manual mapping: implementation sheet "CDIF content item" -> base "CDIF content model"
# or "Generic element" for items that have no CDIF content model value.
impl_to_base_cm = {
    # Clear 1:1 renames
    'Metadata identifier': 'Metadata ID',
    'Metadata date': 'Metadata Date',
    'Metadata contact': 'Metadata Contact Agent',
    'Modification Date': 'Modified Date',
    'Originators': 'Originator',
    'Resource Identifier': 'Resource identifier',
    'Related resources': 'related resource',
    'Distribution/contentURL': 'Distribution.url',
    'Distribution/landingPage': 'URL',
    # Agent variants
    'Other related agents- simple contributor': 'Other related agent',
    'Other related agents - simple contributor': 'Other related agent',
    'related agent with role': 'Other related agent',
}

# Items that map to a base row by Generic element (no CDIF content model)
impl_to_base_ge = {
    'Resource additional type': 'Resource Additional Type',
}

# Items to add as NEW rows (geographic sub-types, SOSO-only, Distribution)
# These get their own rows with CDIF content model populated
new_row_items = {
    'GeographicExtent (bounding box)': 'Geographic Extent (bounding box)',
    'GeographicExtent (named place)': 'Geographic Extent (named place)',
    'GeographicExtent (point location)': 'Geographic Extent (point location)',
    'GeographicExtent (other serialization)': 'Geographic Extent (other serialization)',
    'Distribution': 'Distribution object',  # map to existing CM but allow new row for SOSO/DCAT detail
    'included in data catalog': 'included in data catalog',
    'is accessible for free': 'is accessible for free',
    'isBasedOn': 'isBasedOn',
    'potentialAction': 'potentialAction',
    'prov:used and prov:wasGeneratedBy': 'prov:used and prov:wasGeneratedBy',
    'prov:wasRevisionOf': 'prov:wasRevisionOf',
    'subjectOf': 'subjectOf',
}

# Build lookups
base_by_cm_first = {}
for r in base_rows:
    cm = str(r.get('CDIF content model', '')).strip()
    if cm and cm not in base_by_cm_first:
        base_by_cm_first[cm] = r

# Also build by Generic element for items with no CM
base_by_ge_exact = {}
for r in base_rows:
    ge = str(r.get('Generic element', '')).strip()
    if ge and ge not in base_by_ge_exact:
        base_by_ge_exact[ge] = r

# Track new rows added for new_row_items (by their CDIF content model)
new_impl_rows = {}

def resolve_target(ci):
    """Find or create the target row for a CDIF content item."""
    # Direct CM match
    if ci in base_by_cm_first:
        return base_by_cm_first[ci]
    # Manual CM mapping
    if ci in impl_to_base_cm:
        mapped = impl_to_base_cm[ci]
        if mapped in base_by_cm_first:
            return base_by_cm_first[mapped]
    # Manual GE mapping
    if ci in impl_to_base_ge:
        mapped = impl_to_base_ge[ci]
        if mapped in base_by_ge_exact:
            return base_by_ge_exact[mapped]
    # New row items
    if ci in new_row_items:
        cm_key = new_row_items[ci]
        if cm_key not in new_impl_rows:
            new_row = {h: None for h in all_headers}
            new_row['Generic element'] = ci
            new_row['CDIF content model'] = cm_key
            base_rows.append(new_row)
            new_impl_rows[cm_key] = new_row
        return new_impl_rows[cm_key]
    return None

def apply_impl_row(ci, obl, impl_val, scope_val, soso_comp, source):
    """Apply implementation data to the resolved target row."""
    t = resolve_target(ci)
    if t is None:
        return False
    if obl and not t.get('Obl.'):
        t['Obl.'] = obl
    if source == 'SDO':
        if impl_val:
            t['Schema.org implementation'] = impl_val
        if scope_val:
            t['Scope note (SDO)'] = scope_val
    elif source == 'DCAT':
        if impl_val:
            t['DCAT v3 implementation'] = impl_val
        if scope_val:
            t['Scope note (DCAT)'] = scope_val
    elif source == 'SOSO':
        if impl_val:
            t['Schema.org implementation'] = t.get('Schema.org implementation') or impl_val
        if scope_val:
            t['Scope note (SOSO)'] = scope_val
        if soso_comp:
            t['SOSO comparison'] = soso_comp
    return True

# Schema.org implementation
wb_sdo = openpyxl.load_workbook(
    os.path.join(directory, 'CDIF-schema.org-implementation.xlsx'), data_only=True)
ws_sdo = wb_sdo['Sheet1']
sdo_matched = 0
sdo_unmatched = []
for row in ws_sdo.iter_rows(min_row=2, values_only=True):
    ci = str(row[0]).strip() if row[0] else ''
    if not ci:
        continue
    obl = str(row[1]).strip() if row[1] else None
    impl = str(row[2]).strip() if row[2] else None
    scope = str(row[3]).strip() if row[3] else None
    if apply_impl_row(ci, obl, impl, scope, None, 'SDO'):
        sdo_matched += 1
    else:
        sdo_unmatched.append(ci)
print(f'Schema.org impl matched: {sdo_matched}')
if sdo_unmatched:
    print(f'  Unmatched: {sdo_unmatched}')
wb_sdo.close()

# DCAT implementation
wb_dcat_impl = openpyxl.load_workbook(
    os.path.join(directory, 'CDIF-DCAT-ttl-implementation.xlsx'), data_only=True)
ws_dcat_impl = wb_dcat_impl['Sheet1']
dcat_matched = 0
dcat_unmatched = []
for row in ws_dcat_impl.iter_rows(min_row=2, values_only=True):
    ci = str(row[0]).strip() if row[0] else ''
    if not ci:
        continue
    obl = str(row[1]).strip() if row[1] else None
    impl = str(row[2]).strip() if row[2] else None
    scope = str(row[3]).strip() if row[3] else None
    if apply_impl_row(ci, obl, impl, scope, None, 'DCAT'):
        dcat_matched += 1
    else:
        dcat_unmatched.append(ci)
print(f'DCAT impl matched: {dcat_matched}')
if dcat_unmatched:
    print(f'  Unmatched: {dcat_unmatched}')
wb_dcat_impl.close()

# SOSO comparison
wb_soso = openpyxl.load_workbook(
    os.path.join(directory, 'metadataCrosswalkCDIF-ESIP-SOSO.xlsx'), data_only=True)
ws_soso = wb_soso['Sheet1']
soso_matched = 0
soso_unmatched = []
for row in ws_soso.iter_rows(min_row=2, values_only=True):
    ci = str(row[0]).strip() if row[0] else ''
    if not ci:
        continue
    obl = str(row[1]).strip() if row[1] else None
    impl = str(row[2]).strip() if row[2] else None
    scope = str(row[3]).strip() if row[3] else None
    soso_comp = str(row[4]).strip() if len(row) > 4 and row[4] else None
    if apply_impl_row(ci, obl, impl, scope, soso_comp, 'SOSO'):
        soso_matched += 1
    else:
        soso_unmatched.append(ci)
print(f'SOSO matched: {soso_matched}')
if soso_unmatched:
    print(f'  Unmatched: {soso_unmatched}')
wb_soso.close()

print(f'\nNew rows added from implementation sheets: {len(new_impl_rows)}')
for k in sorted(new_impl_rows.keys()):
    print(f'  {k}')

###############################################################################
# 6. Write merged workbook
###############################################################################
out_wb = openpyxl.Workbook()
out_ws = out_wb.active
out_ws.title = 'Merged Crosswalks'

header_font = Font(bold=True, size=10)
header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
thin_border = Border(bottom=Side(style='thin'))

for col_idx, h in enumerate(all_headers, 1):
    cell = out_ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True)

for row_idx, r in enumerate(base_rows, 2):
    for col_idx, h in enumerate(all_headers, 1):
        val = r.get(h)
        if val is not None:
            out_ws.cell(row=row_idx, column=col_idx, value=val)

for col_idx in range(1, min(9, len(all_headers) + 1)):
    out_ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

out_ws.freeze_panes = 'C2'

# Sources sheet
src_ws = out_wb.create_sheet('Sources')
src_ws.cell(row=1, column=1, value='Source File').font = Font(bold=True)
src_ws.cell(row=1, column=2, value='Sheet').font = Font(bold=True)
src_ws.cell(row=1, column=3, value='Contribution').font = Font(bold=True)
sources = [
    ('CDIF metadata cross walks.xlsx', 'crosswalks draft',
     f'Base table ({len(base_rows)} rows, {len(base_headers)} standard columns)'),
    ('metadataCrosswalkCDIF-DDE-software-etc.xlsx', 'crosswalk',
     f'Added {dde_added} rows, filled empty cells for matching rows'),
    ('metadataCrosswalksCDIF-DCAT-ISOetc.xlsx', 'crosswalks',
     f'Filled {dcat_filled} cells, added {dcat_new} new rows'),
    ('CDIF-schema.org-implementation.xlsx', 'Sheet1',
     f'Added Schema.org implementation + scope notes ({sdo_matched} matches)'),
    ('CDIF-DCAT-ttl-implementation.xlsx', 'Sheet1',
     f'Added DCAT v3 implementation + scope notes ({dcat_matched} matches)'),
    ('metadataCrosswalkCDIF-ESIP-SOSO.xlsx', 'Sheet1',
     f'Added SOSO comparison + scope notes ({soso_matched} matches)'),
    ('CDIF-SOSOCompare.xlsx', 'Sheet1',
     'Duplicate of ESIP-SOSO, not merged separately'),
]
for i, (f, s, c) in enumerate(sources, 2):
    src_ws.cell(row=i, column=1, value=f)
    src_ws.cell(row=i, column=2, value=s)
    src_ws.cell(row=i, column=3, value=c)
src_ws.column_dimensions['A'].width = 50
src_ws.column_dimensions['B'].width = 20
src_ws.column_dimensions['C'].width = 60

out_path = os.path.join(directory, 'CDIF-metadata-crosswalks-merged.xlsx')
out_wb.save(out_path)
print(f'\nSaved: {out_path}')
print(f'Total rows: {len(base_rows)}, Total columns: {len(all_headers)}')
