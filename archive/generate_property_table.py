"""Generate a property table from a building block schema.yaml.

For each building block, produces an Excel workbook with:
- A main worksheet listing the BB's own properties
- Additional worksheets for each referenced class ($defs / $ref targets)

Columns: Field Name, Containing Class, Data Type(s), Cardinality,
         Enum/Const Values, Description.
"""
import yaml
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

###############################################################################
# Configuration
###############################################################################
BB_ROOT = r'C:\Users\smrTu\OneDrive\Documents\GithubC\CDIF\metadataBuildingBlocks\_sources'
CROSSWALK = r'C:\Users\smrTu\OneDrive\Documents\GithubC\CDIF\Discovery\CDIF-metadata-crosswalks-merged.xlsx'

crosswalk_desc = {}
crosswalk_content_model = {}


def load_crosswalk():
    """Load crosswalk descriptions and CDIF content model mappings."""
    try:
        wb = openpyxl.load_workbook(CROSSWALK, read_only=True, data_only=True)
    except (PermissionError, FileNotFoundError) as e:
        print(f"WARNING: Cannot open crosswalk ({e}), skipping crosswalk descriptions.")
        return
    ws = wb['Merged Crosswalks']
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = [str(h).strip() if h else '' for h in row]

    desc_idx = headers.index('Description') if 'Description' in headers else None
    prop_idx = headers.index('schema:Property') if 'schema:Property' in headers else None
    cm_idx = headers.index('CDIF content model') if 'CDIF content model' in headers else None

    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[desc_idx]).strip() if desc_idx is not None and row[desc_idx] else None
        prop_raw = str(row[prop_idx]).strip() if prop_idx is not None and row[prop_idx] else None
        cm = str(row[cm_idx]).strip() if cm_idx is not None and row[cm_idx] else None

        if not prop_raw:
            continue

        # The schema:Property column may have multiple values separated by |
        # or compound entries. Normalize to individual property names.
        prop_variants = [p.strip() for p in prop_raw.replace('|', ',').split(',') if p.strip()]

        for prop in prop_variants:
            if desc:
                crosswalk_desc[prop] = desc
                crosswalk_desc[f'schema:{prop}'] = desc
            if cm:
                crosswalk_content_model[prop] = cm
                crosswalk_content_model[f'schema:{prop}'] = cm
                # Also store without common prefixes/suffixes
                bare = prop.replace('(R)', '').strip()
                if bare != prop:
                    crosswalk_content_model[bare] = cm
                    crosswalk_content_model[f'schema:{bare}'] = cm

    wb.close()
    print(f"Loaded {len(crosswalk_desc)} crosswalk descriptions, {len(crosswalk_content_model)} content model mappings")


def load_schema(path):
    """Load a YAML schema file."""
    with open(path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def resolve_ref(ref, schema, schema_dir):
    """Resolve a $ref and return (resolved_schema, resolved_dir, class_name).

    Follows chains of $ref (e.g. $defs pointing to external files).
    """
    if ref.startswith('#/$defs/'):
        def_name = ref[len('#/$defs/'):]
        defs = schema.get('$defs', {})
        defn = defs.get(def_name, {})
        if '$ref' in defn:
            return resolve_ref(defn['$ref'], schema, schema_dir)
        return defn, schema_dir, def_name
    elif ref.startswith('http://') or ref.startswith('https://'):
        return None, None, ref.split('/')[-2]
    else:
        # Relative file path - strip fragment if present
        frag = None
        if '#' in ref:
            ref, frag = ref.split('#', 1)
        ref_path = os.path.normpath(os.path.join(schema_dir, ref))
        if os.path.exists(ref_path):
            resolved = load_schema(ref_path)
            ref_dir = os.path.dirname(ref_path)
            name = os.path.basename(os.path.dirname(ref_path))
            if frag and frag.startswith('/$defs/'):
                def_name = frag[len('/$defs/'):]
                resolved = resolved.get('$defs', {}).get(def_name, resolved)
                name = def_name
            return resolved, ref_dir, name
        return None, None, ref


def extract_enum_const(prop_schema, schema, schema_dir):
    """Extract enum values and const values from a property schema.

    Returns a string listing the values, or empty string.
    """
    parts = []

    if prop_schema is None:
        return ''

    # Direct const
    if 'const' in prop_schema:
        c = prop_schema['const']
        if isinstance(c, list):
            parts.append('const: ' + ', '.join(str(v) for v in c))
        else:
            parts.append(f'const: {c}')

    # Direct enum
    if 'enum' in prop_schema:
        parts.append(', '.join(str(v) for v in prop_schema['enum']))

    # Default value
    if 'default' in prop_schema and 'const' not in prop_schema and 'enum' not in prop_schema:
        parts.append(f'default: {prop_schema["default"]}')

    # contains (common in @type arrays)
    if 'contains' in prop_schema:
        c = prop_schema['contains']
        if 'const' in c:
            parts.append(f'must contain: {c["const"]}')
        if 'enum' in c:
            parts.append('must contain one of: ' + ', '.join(str(v) for v in c['enum']))

    # Items-level enum/const (for arrays)
    if 'items' in prop_schema:
        items = prop_schema['items']
        if isinstance(items, dict):
            sub = extract_enum_const(items, schema, schema_dir)
            if sub and sub not in '\n'.join(parts):
                parts.append(sub)

    # anyOf / oneOf - collect enum/const from branches
    for key in ('anyOf', 'oneOf'):
        if key in prop_schema:
            for opt in prop_schema[key]:
                sub = extract_enum_const(opt, schema, schema_dir)
                if sub and sub not in '\n'.join(parts):
                    parts.append(sub)

    # allOf
    if 'allOf' in prop_schema:
        for opt in prop_schema['allOf']:
            sub = extract_enum_const(opt, schema, schema_dir)
            if sub and sub not in '\n'.join(parts):
                parts.append(sub)

    # Nested properties - list const/enum/default for all sub-properties
    if 'properties' in prop_schema:
        for sub_name, sub_schema in prop_schema['properties'].items():
            if not isinstance(sub_schema, dict):
                continue
            sub = extract_enum_const(sub_schema, schema, schema_dir)
            if sub:
                parts.append(f'{sub_name}: {sub}')

    return '\n'.join(parts) if parts else ''


def describe_type(prop_schema, schema, schema_dir):
    """Extract a human-readable type description from a property schema."""
    if prop_schema is None:
        return 'any'

    if not isinstance(prop_schema, dict):
        return str(prop_schema)

    if '$ref' in prop_schema:
        _, _, name = resolve_ref(prop_schema['$ref'], schema, schema_dir)
        return name

    if 'anyOf' in prop_schema:
        parts = [describe_type(opt, schema, schema_dir) for opt in prop_schema['anyOf']]
        return ' | '.join(parts)

    if 'oneOf' in prop_schema:
        parts = [describe_type(opt, schema, schema_dir) for opt in prop_schema['oneOf']]
        return ' | '.join(parts)

    if 'allOf' in prop_schema:
        parts = [describe_type(opt, schema, schema_dir) for opt in prop_schema['allOf']]
        return ' & '.join(parts)

    if 'const' in prop_schema:
        return 'const'

    t = prop_schema.get('type')
    if t == 'array':
        items = prop_schema.get('items', {})
        item_type = describe_type(items, schema, schema_dir)
        return f'array of {item_type}'
    elif t == 'object':
        props = prop_schema.get('properties', {})
        if list(props.keys()) == ['@id']:
            return 'object reference'
        # JSON-LD @list pattern: object with single @list property (array)
        if list(props.keys()) == ['@list'] and isinstance(props.get('@list'), dict):
            list_schema = props['@list']
            if list_schema.get('type') == 'array' and 'items' in list_schema:
                item_type = describe_type(list_schema['items'], schema, schema_dir)
                return f'list of {item_type}'
        return 'object'
    elif t == 'string':
        fmt = prop_schema.get('format')
        return f'string ({fmt})' if fmt else 'string'
    elif t == 'number':
        return 'number'
    elif t == 'integer':
        return 'integer'
    elif t == 'boolean':
        return 'boolean'
    elif t:
        return str(t)

    return 'any'


def get_cardinality(field_name, prop_schema, required_fields, any_of_groups):
    """Determine cardinality from schema constraints."""
    is_required = field_name in required_fields

    in_anyof_others = None
    for group in any_of_groups:
        if field_name in group:
            in_anyof_others = [f for f in group if f != field_name]

    t = prop_schema.get('type') if isinstance(prop_schema, dict) else None

    if t == 'array':
        min_items = prop_schema.get('minItems', 0)
        max_items = prop_schema.get('maxItems')
        if is_required:
            low = max(1, min_items)
        else:
            low = 0
        high = str(max_items) if max_items else '*'
        card = f'{low}..{high}'
    else:
        if is_required:
            card = '1'
        else:
            card = '0..1'

    if in_anyof_others is not None and not is_required:
        card += f' (required if no {", ".join(in_anyof_others)})'

    return card


def get_description(field_name, prop_schema):
    """Get description from schema or crosswalk."""
    desc = None
    if isinstance(prop_schema, dict):
        desc = prop_schema.get('description')
    if not desc:
        desc = crosswalk_desc.get(field_name)
    if not desc:
        bare = field_name.replace('schema:', '').replace('dcterms:', '').replace('spdx:', '')
        desc = crosswalk_desc.get(bare)
    return desc or ''


def get_content_model(field_name):
    """Look up CDIF content model from crosswalk by property name."""
    cm = crosswalk_content_model.get(field_name)
    if cm:
        return cm
    # Try without namespace prefix
    bare = field_name.replace('schema:', '').replace('dcterms:', '').replace('spdx:', '')
    cm = crosswalk_content_model.get(bare)
    if cm:
        return cm
    return ''


def get_containing_class(schema):
    """Determine the class name from schema @type or title."""
    props = schema.get('properties', {})
    type_prop = props.get('@type', {})

    # Check contains const
    if 'contains' in type_prop:
        return str(type_prop['contains'].get('const', ''))

    # Check anyOf for const
    for key in ('anyOf', 'oneOf'):
        if key in type_prop:
            for opt in type_prop[key]:
                if 'const' in opt:
                    return str(opt['const'])

    # Check default
    if 'default' in type_prop:
        return str(type_prop['default'])

    # Fall back to title or description
    if schema.get('title'):
        return schema['title']
    if schema.get('description'):
        return schema['description'][:50]

    return 'object'


def collect_required_and_anyof(schema):
    """Collect required fields and anyOf groups from schema."""
    required_fields = set(schema.get('required', []))
    any_of_groups = []

    for item in schema.get('allOf', []):
        if 'required' in item:
            required_fields.update(item['required'])
        if 'anyOf' in item:
            group = []
            for option in item['anyOf']:
                if 'required' in option:
                    group.extend(option['required'])
            if group:
                any_of_groups.append(group)

    return required_fields, any_of_groups


def extract_properties(schema, schema_dir, containing_class=None):
    """Extract property rows from a schema. Returns (rows, referenced_schemas).

    referenced_schemas is a list of (name, resolved_schema, resolved_dir)
    for classes that should get their own worksheets.
    """
    rows = []
    referenced = []

    # For array-type schemas, extract properties from items
    effective_schema = schema
    if schema.get('type') == 'array' and 'items' in schema:
        items = schema['items']
        if isinstance(items, dict) and items.get('properties'):
            effective_schema = items

    if containing_class is None:
        containing_class = get_containing_class(effective_schema)

    required_fields, any_of_groups = collect_required_and_anyof(effective_schema)

    # Merge properties and $defs from top-level and allOf branches.
    # For profiles, allOf entries are composing BBs — their properties are merged
    # into the main sheet and their $defs are merged so fragment refs resolve.
    properties = dict(effective_schema.get('properties', {}))
    merged_defs = dict(effective_schema.get('$defs', {}))
    composing_bb_names = set()   # track composing BBs to exclude from ref sheets
    composing_bb_dirs = {}       # defName -> schema_dir for resolving their $defs
    for item in effective_schema.get('allOf', []):
        if isinstance(item, dict) and item.get('properties'):
            for k, v in item['properties'].items():
                if k not in properties:
                    properties[k] = v
        # Follow $ref in allOf to pull in composed properties
        if isinstance(item, dict) and '$ref' in item:
            resolved, rdir, rname = resolve_ref(item['$ref'], effective_schema, schema_dir)
            if resolved:
                for k, v in resolved.get('properties', {}).items():
                    if k not in properties:
                        properties[k] = v
                # Merge $defs from composing BB so fragment refs can resolve.
                # Resolve each def's $ref chain using the BB's directory so that
                # relative paths resolve correctly regardless of profile location.
                for def_name, def_schema in resolved.get('$defs', {}).items():
                    if def_name not in merged_defs:
                        if isinstance(def_schema, dict) and '$ref' in def_schema and rdir:
                            # Follow the $ref to get the actual schema
                            def_resolved, def_dir, _ = resolve_ref(
                                def_schema['$ref'], resolved, rdir)
                            if def_resolved:
                                merged_defs[def_name] = def_resolved
                                if def_dir:
                                    composing_bb_dirs[def_name] = def_dir
                                continue
                        merged_defs[def_name] = def_schema
                        if rdir:
                            composing_bb_dirs[def_name] = rdir
                # Track composing BB name to skip its own worksheet
                if not item['$ref'].startswith('#'):
                    composing_bb_names.add(rname)
                # Collect required from resolved
                for rr in resolved.get('required', []):
                    required_fields.add(rr)
                for sub in resolved.get('allOf', []):
                    if isinstance(sub, dict) and sub.get('required'):
                        required_fields.update(sub['required'])
                    if isinstance(sub, dict) and sub.get('properties'):
                        for k, v in sub['properties'].items():
                            if k not in properties:
                                properties[k] = v
    # Build an augmented schema with merged $defs for ref resolution
    augmented_schema = dict(schema)
    if merged_defs:
        augmented_schema['$defs'] = merged_defs
    for field_name, prop_schema in properties.items():
        type_desc = describe_type(prop_schema, augmented_schema, schema_dir)
        cardinality = get_cardinality(field_name, prop_schema, required_fields, any_of_groups)
        description = get_description(field_name, prop_schema)
        enum_const = extract_enum_const(prop_schema, augmented_schema, schema_dir)

        content_model = get_content_model(field_name)

        rows.append({
            'Field Name': field_name,
            'Containing Class': containing_class,
            'CDIF Content Model': content_model,
            'Data Type(s)': type_desc,
            'Cardinality': cardinality,
            'Enum/Const Values': enum_const,
            'Description': description,
        })

        # Collect $ref targets for separate worksheets
        collect_refs_from_prop(prop_schema, augmented_schema, schema_dir, referenced)

    # Filter out composing BBs from referenced list — their properties are
    # already merged into the main sheet, so they don't need separate worksheets.
    if composing_bb_names:
        referenced = [(n, s, d) for n, s, d in referenced
                      if n not in composing_bb_names]

    return rows, referenced


def has_properties(schema_obj):
    """Check if a schema defines properties (directly, in items, or in allOf)."""
    if not isinstance(schema_obj, dict):
        return False
    if schema_obj.get('properties'):
        return True
    # Array type with object items that have properties
    if schema_obj.get('type') == 'array':
        items = schema_obj.get('items', {})
        if isinstance(items, dict) and items.get('properties'):
            return True
    # allOf composition — check if any branch has properties
    for item in schema_obj.get('allOf', []):
        if isinstance(item, dict) and item.get('properties'):
            return True
    return False


def collect_refs_from_prop(prop_schema, schema, schema_dir, referenced):
    """Recursively collect $ref targets from a property schema."""
    if not isinstance(prop_schema, dict):
        return

    if '$ref' in prop_schema:
        resolved, rdir, name = resolve_ref(prop_schema['$ref'], schema, schema_dir)
        if resolved and rdir and has_properties(resolved):
            referenced.append((name, resolved, rdir))
        return

    # Check items
    if 'items' in prop_schema:
        items = prop_schema['items']
        if isinstance(items, dict):
            collect_refs_from_prop(items, schema, schema_dir, referenced)

    # Check anyOf/oneOf/allOf branches
    for key in ('anyOf', 'oneOf', 'allOf'):
        if key in prop_schema:
            for opt in prop_schema[key]:
                collect_refs_from_prop(opt, schema, schema_dir, referenced)


def write_workbook(main_rows, ref_sheets, output_path, main_sheet_name):
    """Write property table workbook with multiple worksheets."""
    wb = openpyxl.Workbook()
    headers = ['Field Name', 'Containing Class', 'CDIF Content Model',
               'Data Type(s)', 'Cardinality', 'Enum/Const Values', 'Description']
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    def write_sheet(ws, rows):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True)

        for row_idx, r in enumerate(rows, 2):
            for col_idx, h in enumerate(headers, 1):
                val = r.get(h, '')
                if val:
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 80
        ws.freeze_panes = 'A2'

    # Main sheet
    main_ws = wb.active
    main_ws.title = main_sheet_name[:31]  # Excel 31-char limit
    write_sheet(main_ws, main_rows)

    # Referenced class sheets
    seen_names = {main_sheet_name}
    for name, rows in ref_sheets:
        if name in seen_names:
            continue
        seen_names.add(name)
        sheet_name = name[:31]
        ws = wb.create_sheet(title=sheet_name)
        write_sheet(ws, rows)

    wb.save(output_path)
    total = len(main_rows) + sum(len(rows) for _, rows in ref_sheets)
    print(f'Saved: {output_path}')
    print(f'  Main sheet: {main_sheet_name} ({len(main_rows)} properties)')
    for name, rows in ref_sheets:
        if name in seen_names:
            print(f'  Sheet: {name} ({len(rows)} properties)')
            seen_names.discard(name)  # only print once


def process_building_block(bb_path):
    """Process a building block and generate its property table workbook."""
    schema_dir = os.path.dirname(bb_path)
    schema = load_schema(bb_path)
    bb_name = os.path.basename(schema_dir)

    print(f'\nProcessing: {bb_path}')
    print(f'Title: {schema.get("title", schema.get("description", "N/A")[:80])}')

    main_rows, referenced = extract_properties(schema, schema_dir)

    # Process referenced schemas recursively into their own sheets
    ref_sheets = []
    seen = set()
    queue = list(referenced)
    while queue:
        name, ref_schema, ref_dir = queue.pop(0)
        if name in seen:
            continue
        seen.add(name)
        ref_rows, sub_refs = extract_properties(ref_schema, ref_dir)
        if ref_rows:
            ref_sheets.append((name, ref_rows))
            # Add sub-refs to queue for processing
            queue.extend(sub_refs)

    # Console output
    def print_rows(rows, label=None):
        if label:
            print(f'\n--- {label} ---')
        for r in rows:
            desc = r['Description'][:40] + '...' if len(r['Description']) > 40 else r['Description']
            cm = r.get('CDIF Content Model', '')
            print(f'{r["Field Name"]:28s} {r["Containing Class"]:22s} {cm:25s} {r["Data Type(s)"]:28s} {r["Cardinality"]:12s} {desc}')

    print(f'\n{"Field Name":28s} {"Class":22s} {"CDIF Content Model":25s} {"Type":28s} {"Card":12s} Description')
    print('-' * 170)
    print_rows(main_rows)
    for name, rows in ref_sheets:
        print_rows(rows, name)

    # Write Excel
    output_path = os.path.join(schema_dir, f'{bb_name}_properties.xlsx')
    write_workbook(main_rows, ref_sheets, output_path, bb_name)


def main():
    load_crosswalk()

    bb_path = os.path.join(BB_ROOT, 'cdifProperties', 'cdifMandatory', 'schema.yaml')
    if len(sys.argv) > 1:
        bb_path = sys.argv[1]

    process_building_block(bb_path)


if __name__ == '__main__':
    main()
